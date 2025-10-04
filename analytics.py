import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import text
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

from config import ENGINE, CHARTS_DIR, EXPORTS_DIR

CHARTS_DIR.mkdir(parents=True, exist_ok=True)
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

def run_query(sql):
    with ENGINE.connect() as conn:
        df = pd.read_sql(text(sql), conn)
        print(f"\n[SQL EXECUTED]\n{sql.strip()}")
        print(df.head(), f"\nRows: {len(df)}")
        return df


def pie_orders_by_state():
    df = run_query("""
        SELECT c.customer_state, COUNT(o.order_id) AS orders_count
        FROM olist_customers_dataset c
        JOIN olist_orders_dataset o ON c.customer_id = o.customer_id
        GROUP BY c.customer_state
        ORDER BY orders_count DESC
        LIMIT 10;
    """)
    plt.pie(df['orders_count'], labels=df['customer_state'], autopct='%1.1f%%')
    plt.title("Orders by State (Top 10)")
    plt.savefig(CHARTS_DIR / "pie.png")
    plt.close()

def bar_orders_by_city():
    df = run_query("""
        SELECT c.customer_city, COUNT(o.order_id) AS orders_count
        FROM olist_customers_dataset c
        JOIN olist_orders_dataset o ON c.customer_id = o.customer_id
        GROUP BY c.customer_city
        ORDER BY orders_count DESC
        LIMIT 20;
    """)
    plt.bar(df['customer_city'], df['orders_count'])
    plt.xticks(rotation=45, ha='right')
    plt.title("Orders by City (Top 20)")
    plt.savefig(CHARTS_DIR / "bar.png")
    plt.close()

def hbar_avg_delay():
    df = run_query("""
        SELECT c.customer_state,
               AVG(EXTRACT(EPOCH FROM (o.order_delivered_customer_date::timestamp - o.order_estimated_delivery_date::timestamp))/86400.0) AS avg_delay
        FROM olist_customers_dataset c
        JOIN olist_orders_dataset o ON c.customer_id = o.customer_id
        WHERE o.order_delivered_customer_date IS NOT NULL AND o.order_estimated_delivery_date IS NOT NULL
        GROUP BY c.customer_state
        ORDER BY avg_delay DESC
        LIMIT 15;
    """)
    df.set_index('customer_state')['avg_delay'].plot.barh()
    plt.title("Average Delivery Delay by State")
    plt.savefig(CHARTS_DIR / "hbar.png")
    plt.close()

def line_orders_over_time():
    df = run_query("""
        SELECT DATE(order_purchase_timestamp) AS order_date, COUNT(*) AS orders_count
        FROM olist_orders_dataset
        WHERE order_purchase_timestamp IS NOT NULL
        GROUP BY DATE(order_purchase_timestamp)
        ORDER BY order_date;
    """)
    plt.plot(df['order_date'], df['orders_count'])
    plt.title("Orders Over Time")
    plt.savefig(CHARTS_DIR / "line.png")
    plt.close()

def hist_orders_per_customer():
    df = run_query("""
        SELECT c.customer_id, COUNT(o.order_id) AS orders_per_customer
        FROM olist_customers_dataset c
        JOIN olist_orders_dataset o ON c.customer_id = o.customer_id
        GROUP BY c.customer_id;
    """)
    plt.hist(df['orders_per_customer'], bins=20)
    plt.title("Histogram of Orders per Customer")
    plt.savefig(CHARTS_DIR / "hist.png")
    plt.close()

def scatter_orders_vs_score():
    df = run_query("""
        SELECT c.customer_id, COUNT(o.order_id) AS orders_count, AVG(r.review_score::numeric) AS avg_score
        FROM olist_customers_dataset c
        JOIN olist_orders_dataset o ON c.customer_id = o.customer_id
        JOIN olist_order_reviews_dataset r ON o.order_id = r.order_id
        GROUP BY c.customer_id
        LIMIT 500;
    """)
    plt.scatter(df['orders_count'], df['avg_score'])
    plt.title("Orders vs Review Score")
    plt.xlabel("Orders Count")
    plt.ylabel("Average Score")
    plt.savefig(CHARTS_DIR / "scatter.png")
    plt.close()

# ---------- Plotly Slider ----------
def plotly_slider():
    df = run_query("""
        SELECT DATE(order_purchase_timestamp) AS order_date, order_id
        FROM olist_orders_dataset
        WHERE order_purchase_timestamp IS NOT NULL;
    """)
    df['month'] = pd.to_datetime(df['order_date']).dt.to_period("M").astype(str)
    grouped = df.groupby('month').order_id.count().reset_index(name="orders_count")
    fig = px.bar(grouped, x="month", y="orders_count", animation_frame="month",
                 title="Orders by Month")
    fig.write_html(CHARTS_DIR / "plotly_slider.html")

# ---------- Excel Export ----------
def export_excel():
    customers = run_query("SELECT * FROM olist_customers_dataset LIMIT 500;")
    orders = run_query("SELECT * FROM olist_orders_dataset LIMIT 500;")
    path = EXPORTS_DIR / "report.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        customers.to_excel(writer, sheet_name="customers", index=False)
        orders.to_excel(writer, sheet_name="orders", index=False)
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        max_col = ws.max_column
        max_row = ws.max_row
        for col in range(2, max_col+1):
            col_letter = ws.cell(row=1, column=col).column_letter
            rng = f"{col_letter}2:{col_letter}{max_row}"
            rule = ColorScaleRule(start_type="min", start_color="FFAAAAFF",
                                  mid_type="percentile", mid_value=50, mid_color="FFFFFFAA",
                                  end_type="max", end_color="FFAAFFAA")
            ws.conditional_formatting.add(rng, rule)
    wb.save(path)


def main():
    pie_orders_by_state()
    bar_orders_by_city()
    hbar_avg_delay()
    line_orders_over_time()
    hist_orders_per_customer()
    scatter_orders_vs_score()
    plotly_slider()
    export_excel()

if __name__ == "__main__":
    main()
